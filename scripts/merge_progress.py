"""Merge user-edited fields from an existing checklist workbook into a newly generated one."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def _cell_key(*parts: str) -> str:
    return "||".join(str(p).strip() for p in parts if p is not None)


def _read_sheet_map(ws, key_cols: list[int], value_cols: list[int], start_row: int = 2) -> dict[str, dict[int, object]]:
    out: dict[str, dict[int, object]] = {}
    for r in range(start_row, ws.max_row + 1):
        key_parts = [ws.cell(r, c).value for c in key_cols]
        if not any(key_parts):
            continue
        key = _cell_key(*key_parts)
        vals = {c: ws.cell(r, c).value for c in value_cols}
        if any(v not in (None, "") for v in vals.values()):
            out[key] = vals
    return out


def _apply_map(ws, data: dict[str, dict[int, object]], key_cols: list[int], start_row: int = 2) -> int:
    applied = 0
    for r in range(start_row, ws.max_row + 1):
        key_parts = [ws.cell(r, c).value for c in key_cols]
        if not any(key_parts):
            continue
        key = _cell_key(*key_parts)
        if key not in data:
            continue
        for col, val in data[key].items():
            if val not in (None, ""):
                ws.cell(r, col, val)
                applied += 1
    return applied


def merge_workbook(new_path: Path, old_path: Path) -> dict[str, int]:
    """Apply progress from old_path into new_path (in place). Returns per-sheet apply counts."""
    old_wb = load_workbook(old_path, data_only=True)
    new_wb = load_workbook(new_path)
    stats: dict[str, int] = {}

    if "01_项目总览" in old_wb.sheetnames and "01_项目总览" in new_wb.sheetnames:
        old_ws, new_ws = old_wb["01_项目总览"], new_wb["01_项目总览"]
        m = _read_sheet_map(old_ws, [1], [11, 13, 14, 15, 22, 23, 24, 25, 26])
        stats["01_项目总览"] = _apply_map(new_ws, m, [1])

    if "02_材料进度" in old_wb.sheetnames and "02_材料进度" in new_wb.sheetnames:
        old_ws, new_ws = old_wb["02_材料进度"], new_wb["02_材料进度"]
        m = _read_sheet_map(old_ws, [1, 3], [5, 7, 10, 12], start_row=3)
        stats["02_材料进度"] = _apply_map(new_ws, m, [1, 3], start_row=3)

    if "03_推荐信进度" in old_wb.sheetnames and "03_推荐信进度" in new_wb.sheetnames:
        old_ws, new_ws = old_wb["03_推荐信进度"], new_wb["03_推荐信进度"]
        for r in range(3, 6):
            for c in range(2, 8):
                val = old_ws.cell(r, c).value
                if val not in (None, ""):
                    new_ws.cell(r, c, val)
        m = _read_sheet_map(old_ws, [1, 4], [6, 7, 8, 9, 10, 11, 12], start_row=9)
        stats["03_推荐信进度"] = _apply_map(new_ws, m, [1, 4], start_row=9)

    if "04_Essay进度" in old_wb.sheetnames and "04_Essay进度" in new_wb.sheetnames:
        old_ws, new_ws = old_wb["04_Essay进度"], new_wb["04_Essay进度"]
        m = _read_sheet_map(old_ws, [1, 3], [5, 6, 7, 9, 11, 12, 13])
        stats["04_Essay进度"] = _apply_map(new_ws, m, [1, 3])

    if "05_网申账号" in old_wb.sheetnames and "05_网申账号" in new_wb.sheetnames:
        old_ws, new_ws = old_wb["05_网申账号"], new_wb["05_网申账号"]
        m = _read_sheet_map(old_ws, [1], [6, 7, 9, 10, 11])
        stats["05_网申账号"] = _apply_map(new_ws, m, [1])

    if "06_录取进度" in old_wb.sheetnames and "06_录取进度" in new_wb.sheetnames:
        old_ws, new_ws = old_wb["06_录取进度"], new_wb["06_录取进度"]
        m = _read_sheet_map(old_ws, [1], list(range(4, 13)))
        stats["06_录取进度"] = _apply_map(new_ws, m, [1])

    new_wb.save(new_path)
    return stats
