#!/usr/bin/env python3
"""Generate application progress checklist workbook from config + school list CSV."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))
from merge_progress import merge_workbook  # noqa: E402

SHARED_PROJECT_ID = "通用"

APP_STATUS = ["未注册", "已注册", "填写中", "待支付", "已提交", "Complete"]
MATERIAL_STATUS = [
    "未开始", "准备中", "待审核", "已定稿", "已上传", "已送分/Official", "不适用",
]
LOR_STATUS = ["未邀请", "已邀请", "已提醒", "已提交", "逾期", "不适用"]
ESSAY_VERSION = ["V0大纲", "V1草稿", "V2修订", "V3定稿", "已上传"]
ESSAY_STATUS = [
    "未开始", "大纲中", "写作中", "待学生反馈", "待顾问修改", "已定稿", "已上传",
]
ADMISSION_STATUS = [
    "已提交", "Under Review", "Interview", "Decision Pending",
    "Admitted", "Waitlisted", "Rejected", "Deferred",
]
OWNER = ["顾问", "学生", "双方"]
PAY_STATUS = ["未支付", "已支付", "已退款"]
REG_STATUS = ["未注册", "已注册"]
YES_NO = ["是", "否"]
REQUIRED = ["是", "否", "Optional"]

DEFAULT_SHARED_MATERIALS = [
    ("UW 成绩单", "身份/学历", "学生", "Documents", "是", ""),
    ("OSU 成绩单", "身份/学历", "学生", "Documents", "是", ""),
    ("学位证明/在读证明", "身份/学历", "学生", "Documents", "是", ""),
    ("TOEFL/IELTS 送分", "标化", "学生", "Test Scores", "是", ""),
    ("Resume/CV", "文书类", "双方", "Resume", "是", ""),
]

DEFAULT_PROJECT_MATERIALS = [
    ("网申表填写", "网申填写", "学生", "Application Form", "是"),
    ("先修课程/Background 说明", "网申填写", "双方", "Background", "是"),
    ("申请费支付", "费用", "学生", "Payment", "是"),
]

SCHOOL_ABBR = {
    "university of texas": "UT",
    "texas at austin": "UT",
    "university of pennsylvania": "PENN",
    "pennsylvania": "PENN",
    "georgia institute of technology": "GT",
    "georgia tech": "GT",
    "carnegie mellon": "CMU",
    "university of illinois": "UIUC",
    "illinois urbana": "UIUC",
    "stanford": "STAN",
    "columbia": "COL",
    "usc": "USC",
    "johns hopkins": "JHU",
    "northeastern": "NEU",
    "arizona state": "ASU",
}

PROGRAM_ABBR_RULES = [
    (r"artificial intelligence|msai|\bai\b", "MSAI"),
    (r"software engineering|msse|ms-se", "MSSE"),
    (r"omscs|online master.*computer science", "OMSCS"),
    (r"computer science online|mcs online|online mcs", "MCS"),
    (r"mse-ai|mse ai", "MSEAI"),
    (r"master.*computer science|mscs|\bmscs\b", "MSCS"),
    (r"data science", "MSDS"),
    (r"information networking|msin", "MSIN"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")


@dataclass
class ProjectMeta:
    lor_count: int = 3
    lor_optional_third: bool = False
    fee: int | None = None
    portal_system: str = ""
    portal_url: str = ""
    local_dir: str = ""
    essays: list[tuple[str, str, str]] = field(default_factory=list)
    account_note: str = ""


@dataclass
class AppConfig:
    student_name: str
    advisor_name: str
    output_dir: Path
    school_list_csv: Path
    research_csv: Path | None
    output_filename: str
    project_id_map: dict[str, str]
    school_en_fixes: dict[str, str]
    projects: dict[str, dict]
    shared_materials: list[tuple]
    project_materials: list[tuple]
    readiness_weights: dict[str, dict[str, float]]
    resume_local_path: str


def parse_date(value: str) -> datetime | str:
    value = (value or "").strip()
    if not value:
        return ""
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return value


def abbreviate_school(school_en: str) -> str:
    lower = school_en.lower()
    for key, abbr in SCHOOL_ABBR.items():
        if key in lower:
            return abbr
    words = re.findall(r"[A-Za-z]+", school_en)
    if not words:
        return "SCH"
    if len(words) >= 2:
        return (words[0][:2] + words[1][:2]).upper()
    return words[0][:4].upper()


def abbreviate_program(program: str) -> str:
    lower = program.lower()
    for pattern, abbr in PROGRAM_ABBR_RULES:
        if re.search(pattern, lower):
            return abbr
    words = re.findall(r"[A-Za-z]+", program)
    return "".join(w[:3] for w in words[:2]).upper() or "PROG"


def make_project_id(program: str, school_en: str, existing: set[str], id_map: dict[str, str]) -> str:
    if program in id_map:
        return id_map[program]
    base = f"{abbreviate_school(school_en)}-{abbreviate_program(program)}"
    pid = base
    n = 2
    while pid in existing:
        pid = f"{base}{n}"
        n += 1
    existing.add(pid)
    return pid


def parse_fee(raw: str) -> int | None:
    if not raw:
        return None
    nums = re.findall(r"\d+", raw.replace(",", ""))
    return int(nums[-1]) if nums else None


def parse_lor_count(raw: str) -> tuple[int, bool]:
    if not raw:
        return 3, False
    raw = raw.strip()
    if "不要求" in raw or raw.startswith("0"):
        return 0, False
    nums = re.findall(r"\d+", raw)
    count = int(nums[0]) if nums else 3
    optional_third = "可选" in raw or "optional" in raw.lower()
    if optional_third and count <= 2:
        count = 3
    return count, optional_third


def infer_essays(program: str, materials_text: str, local_dir: str) -> list[tuple[str, str, str]]:
    lower = (program + " " + materials_text).lower()
    essays: list[tuple[str, str, str]] = []
    if "academic sop" in lower or ("personal statement" in lower and "sop" in lower):
        essays.append(("Academic SOP", "按官网", local_dir))
        essays.append(("Personal Statement", "按官网", local_dir))
    elif "personal statement" in lower:
        essays.append(("Personal Statement", "按官网", local_dir))
    elif "statement of purpose" in lower or "sop" in lower:
        essays.append(("SOP", "按官网", local_dir))
    else:
        essays.append(("SOP", "按官网", local_dir))
    if "short-answer" in lower or "short answer" in lower:
        essays.append(("Short Answers", "按官网", local_dir))
    return essays


def load_config(config_path: Path) -> AppConfig:
    with config_path.open(encoding="utf-8") as f:
        raw = yaml.safe_load(f)
    base = config_path.parent
    return AppConfig(
        student_name=raw["student_name"],
        advisor_name=raw.get("advisor_name", ""),
        output_dir=(base / raw.get("output_dir", ".")).resolve(),
        school_list_csv=(base / raw["school_list_csv"]).resolve(),
        research_csv=(base / raw["research_csv"]).resolve() if raw.get("research_csv") else None,
        output_filename=raw.get("output_filename", f"{raw['student_name']}_申请进度Checklist.xlsx"),
        project_id_map=raw.get("project_id_map", {}),
        school_en_fixes=raw.get("school_en_fixes", {}),
        projects=raw.get("projects", {}),
        shared_materials=_parse_shared_materials(raw.get("shared_materials")),
        project_materials=_parse_project_materials(raw.get("project_materials")),
        readiness_weights=raw.get(
            "readiness_weights",
            {
                "default": {"material": 0.4, "lor": 0.3, "essay": 0.2, "app": 0.1},
                "no_lor": {"material": 0.4, "lor": 0.0, "essay": 0.2, "app": 0.1},
            },
        ),
        resume_local_path=raw.get("resume_local_path", "文书/source/Resume.md"),
    )


def _parse_shared_materials(items) -> list[tuple]:
    if not items:
        mats = [list(m) for m in DEFAULT_SHARED_MATERIALS]
        for m in mats:
            if m[0] == "Resume/CV":
                m[5] = ""  # filled from config later
        return [tuple(m) for m in mats]
    out = []
    for item in items:
        out.append((
            item["name"],
            item.get("category", "身份/学历"),
            item.get("owner", "学生"),
            item.get("portal", "Documents"),
            item.get("required", "是"),
            item.get("local_path", ""),
        ))
    return out


def _parse_project_materials(items) -> list[tuple]:
    if not items:
        return list(DEFAULT_PROJECT_MATERIALS)
    return [
        (i["name"], i.get("category", "网申填写"), i.get("owner", "学生"), i.get("portal", ""), i.get("required", "是"))
        for i in items
    ]


def load_research_index(research_csv: Path | None) -> dict[str, dict]:
    if not research_csv or not research_csv.exists():
        return {}
    index: dict[str, dict] = {}
    with research_csv.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = (row.get("专业英文名称") or row.get("项目") or "").strip().lower()
            if key:
                index[key] = row
    return index


def match_research_row(program: str, school_en: str, index: dict[str, dict]) -> dict | None:
    prog_lower = program.lower()
    for key, row in index.items():
        if key in prog_lower or prog_lower in key:
            return row
    school_lower = school_en.lower()
    for row in index.values():
        uni = (row.get("大学英文名称") or "").lower()
        pname = (row.get("专业英文名称") or "").lower()
        if uni and uni in school_lower and (not pname or pname in prog_lower or prog_lower in pname):
            return row
    return None


def parse_projects(cfg: AppConfig) -> tuple[list[dict], dict[str, ProjectMeta]]:
    research = load_research_index(cfg.research_csv)
    existing_ids: set[str] = set()
    projects: list[dict] = []
    meta_map: dict[str, ProjectMeta] = {}

    with cfg.school_list_csv.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            program = (row.get("项目") or "").strip()
            if not program:
                continue
            school_en = cfg.school_en_fixes.get(program, (row.get("UniverisityUniversity") or row.get("University") or "").strip())
            pid = make_project_id(program, school_en, existing_ids, cfg.project_id_map)
            ddl = (row.get("截止日期") or "").strip()
            note = (row.get("申请备注/录取条件") or row.get("申请备注") or "").strip()
            final_ddl = ""
            if re.search(r"9/1\s*final|final.*9/1", note, re.I):
                final_ddl = "2026-09-01"

            p = {
                "id": pid,
                "school_zh": (row.get("学校中文名") or "").strip(),
                "school_en": school_en,
                "program": program,
                "link": (row.get("项目链接") or "").strip(),
                "term": (row.get("入学时间") or "").strip(),
                "priority_ddl": parse_date(ddl),
                "final_ddl": parse_date(final_ddl or ddl),
                "note": note,
                "tier": (row.get("申请梯队") or row.get("梯队") or "").strip(),
            }
            projects.append(p)

            research_row = match_research_row(program, school_en, research)
            lor_count, lor_opt = 3, False
            fee = None
            portal_url = p["link"]
            portal_system = p["school_zh"] or school_en

            if research_row:
                lor_count, lor_opt = parse_lor_count(research_row.get("RL数量", ""))
                fee = parse_fee(research_row.get("申请费（美元)", "") or research_row.get("申请费", ""))
                portal_url = (research_row.get("招生网址") or portal_url).strip()
                portal_system = (research_row.get("所属院系") or portal_system).strip()[:40]

            overrides = cfg.projects.get(pid, {})
            if "lor_count" in overrides:
                lor_count = int(overrides["lor_count"])
            if "lor_optional_third" in overrides:
                lor_opt = bool(overrides["lor_optional_third"])
            if "fee" in overrides:
                fee = int(overrides["fee"])
            if overrides.get("portal_url"):
                portal_url = overrides["portal_url"]
            if overrides.get("portal_system"):
                portal_system = overrides["portal_system"]

            local_dir = overrides.get("local_dir") or f"申请/{pid}/"
            materials_text = (research_row or {}).get("材料要求", "")
            if overrides.get("essays"):
                essays = [(e, "按官网", local_dir) for e in overrides["essays"]]
            else:
                essays = infer_essays(program, materials_text, local_dir)

            meta_map[pid] = ProjectMeta(
                lor_count=lor_count,
                lor_optional_third=lor_opt,
                fee=fee or 90,
                portal_system=portal_system,
                portal_url=portal_url,
                local_dir=local_dir,
                essays=essays,
                account_note=overrides.get("account_note", ""),
            )

    return projects, meta_map


def style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row: int, end_row: int, col_count: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_list_validation(ws, cell_range: str, options: list[str]) -> None:
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    dv.error = "请从下拉列表中选择"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_table(ws, ref: str, name: str) -> None:
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)


def build_workbook(cfg: AppConfig, projects: list[dict], meta_map: dict[str, ProjectMeta]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    build_projects_sheet(wb, cfg, projects, meta_map)
    build_materials_sheet(wb, cfg, projects, meta_map)
    build_lor_sheet(wb, projects, meta_map)
    build_essay_sheet(wb, projects, meta_map)
    build_account_sheet(wb, projects, meta_map)
    build_admission_sheet(wb, projects)
    build_dashboard(wb, cfg, projects, meta_map)
    order = ["00_仪表盘", "01_项目总览", "02_材料进度", "03_推荐信进度", "04_Essay进度", "05_网申账号", "06_录取进度"]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
    fix_overall_formula(wb, projects, meta_map, cfg.readiness_weights)
    apply_conditional_formatting(wb, len(projects))
    return wb


def build_projects_sheet(wb, cfg, projects, meta_map):
    ws = wb.create_sheet("01_项目总览")
    headers = [
        "项目ID", "学校中文名", "学校英文名", "项目名称", "项目链接", "入学批次",
        "Priority DDL", "Final DDL", "申请备注", "网申系统", "网申状态", "申请费($)",
        "支付状态", "目标提交日", "实际提交日", "Portal链接", "材料进度", "推荐信进度",
        "Essay进度", "网申完成度(%)", "整体就绪度(%)", "当前卡点", "下一步动作",
        "负责人", "最后更新人", "最后更新日期",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for p in projects:
        m = meta_map[p["id"]]
        ws.append([
            p["id"], p["school_zh"], p["school_en"], p["program"], p["link"], p["term"],
            p["priority_ddl"], p["final_ddl"], p["note"], m.portal_system, "未注册",
            m.fee, "未支付", "", "", m.portal_url, "", "", "", "", "", "", "", "双方", "", "",
        ])
    first, last = 2, len(projects) + 1
    for r in range(first, last + 1):
        pid = ws.cell(r, 1).value
        m = meta_map[pid]
        scope = f"(((('02_材料进度'!$A:$A=$A{r})+('02_材料进度'!$A:$A=\"{SHARED_PROJECT_ID}\"))>0)"
        done = (
            f"SUMPRODUCT({scope}*('02_材料进度'!$F:$F=\"是\")*"
            f"(('02_材料进度'!$E:$E=\"已定稿\")+('02_材料进度'!$E:$E=\"已上传\")+('02_材料进度'!$E:$E=\"已送分/Official\")))"
        )
        total = f"SUMPRODUCT({scope}*('02_材料进度'!$F:$F=\"是\"))"
        ws.cell(r, 17).value = f'=IF({total}=0,"0/0",{done}&"/"&{total})'
        if m.lor_count == 0:
            ws.cell(r, 18).value = "N/A"
        else:
            ws.cell(r, 18).value = (
                f'=IF(SUMPRODUCT((\'03_推荐信进度\'!$A:$A=$A{r})*(\'03_推荐信进度\'!$F:$F="是"))=0,"0/0",'
                f'SUMPRODUCT((\'03_推荐信进度\'!$A:$A=$A{r})*(\'03_推荐信进度\'!$F:$F="是")*(\'03_推荐信进度\'!$G:$G="已提交"))'
                f'&"/"&SUMPRODUCT((\'03_推荐信进度\'!$A:$A=$A{r})*(\'03_推荐信进度\'!$F:$F="是")))'
            )
        ws.cell(r, 19).value = (
            f'=IF(COUNTIF(\'04_Essay进度\'!$A:$A,$A{r})=0,"0/0",'
            f'SUMPRODUCT((\'04_Essay进度\'!$A:$A=$A{r})*((\'04_Essay进度\'!$F:$F="已定稿")+(\'04_Essay进度\'!$F:$F="已上传")))'
            f'&"/"&COUNTIF(\'04_Essay进度\'!$A:$A,$A{r}))'
        )
        ws.cell(r, 20).value = (
            f'=IF($K{r}="Complete",100,IF($K{r}="已提交",90,IF($K{r}="待支付",70,'
            f'IF($K{r}="填写中",40,IF($K{r}="已注册",10,0)))))'
        )
    style_data_rows(ws, first, last, len(headers))
    for r in range(first, last + 1):
        for col in (7, 8, 14, 15, 26):
            cell = ws.cell(r, col)
            if isinstance(cell.value, datetime):
                cell.number_format = "YYYY-MM-DD"
    ws.freeze_panes = "A2"
    add_list_validation(ws, f"K{first}:K{last}", APP_STATUS)
    add_list_validation(ws, f"M{first}:M{last}", PAY_STATUS)
    add_list_validation(ws, f"X{first}:X{last}", OWNER)
    add_table(ws, f"A1:{get_column_letter(len(headers))}{last}", "tblProjects")


def build_materials_sheet(wb, cfg, projects, meta_map):
    ws = wb.create_sheet("02_材料进度")
    ws["A1"] = "说明：项目ID=「通用」的材料所有学校共用，只需维护一次；其余按项目单独跟踪。"
    ws["A1"].font = Font(bold=True, size=11, color="1F4E79")
    ws.merge_cells("A1:L1")
    headers = [
        "项目ID", "适用范围", "材料名称", "材料类别", "状态", "是否必需",
        "本地文件路径", "Portal对应栏目", "截止日期", "完成日期", "负责人", "备注",
    ]
    hr = 2
    for c, h in enumerate(headers, 1):
        ws.cell(hr, c, h)
    style_header_row(ws, hr, len(headers))
    ds = hr + 1
    rows = []
    for name, cat, owner, portal, req, local in cfg.shared_materials:
        if name == "Resume/CV" and not local:
            local = cfg.resume_local_path
        rows.append([SHARED_PROJECT_ID, "全部项目", name, cat, "未开始", req, local, portal, "", "", owner, "所有学校共用，更新一次即可"])
    for p in projects:
        for name, cat, owner, portal, req in cfg.project_materials:
            rows.append([p["id"], "单项目", name, cat, "未开始", req, "", portal, p["priority_ddl"], "", owner, ""])
    for i, row in enumerate(rows):
        for c, val in enumerate(row, 1):
            cell = ws.cell(ds + i, c, val)
            if c == 9 and isinstance(val, datetime):
                cell.number_format = "YYYY-MM-DD"
    last = ds + len(rows) - 1
    style_data_rows(ws, ds, last, len(headers))
    shared_fill = PatternFill("solid", fgColor="E2EFDA")
    for r in range(ds, last + 1):
        if ws.cell(r, 1).value == SHARED_PROJECT_ID:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = shared_fill
    ws.freeze_panes = f"A{ds}"
    add_list_validation(ws, f"E{ds}:E{last}", MATERIAL_STATUS)
    add_list_validation(ws, f"F{ds}:F{last}", REQUIRED)
    add_list_validation(ws, f"K{ds}:K{last}", OWNER)
    add_table(ws, f"A{hr}:{get_column_letter(len(headers))}{last}", "tblMaterials")


def build_lor_sheet(wb, projects, meta_map):
    ws = wb.create_sheet("03_推荐信进度")
    ws["A1"] = "推荐人主数据（请先填写）"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:H1")
    for c, h in enumerate(["编号", "姓名", "职称/关系", "邮箱", "是否已同意", "LoR基础版状态", "备注", ""], 1):
        cell = ws.cell(2, c, h)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
    for i, rid in enumerate(["R1", "R2", "R3"], 3):
        ws.cell(i, 1, rid)
    add_list_validation(ws, "E3:E5", YES_NO)
    add_list_validation(ws, "F3:F5", ["未开始", "准备中", "已定稿", "不适用"])
    ds = 7
    ws.cell(ds, 1, "推荐信分配明细").font = Font(bold=True, size=12)
    headers = ["项目ID", "学校", "项目名称", "推荐人编号", "推荐人姓名", "是否分配", "邀请/提交状态",
               "邀请发送日", "最近提醒日", "提交确认日", "Portal显示状态", "备注"]
    hr = ds + 1
    for c, h in enumerate(headers, 1):
        ws.cell(hr, c, h)
    style_header_row(ws, hr, len(headers))
    cur = hr + 1
    for p in projects:
        m = meta_map[p["id"]]
        if m.lor_count == 0:
            continue
        rids = ["R1", "R2", "R3"] if m.lor_count >= 3 else ["R1", "R2"]
        for rid in rids:
            assign = "否" if m.lor_optional_third and rid == "R3" else "是"
            ws.cell(cur, 1, p["id"])
            ws.cell(cur, 2, p["school_zh"])
            ws.cell(cur, 3, p["program"])
            ws.cell(cur, 4, rid)
            ws.cell(cur, 5).value = f'=IF(D{cur}="R1",$B$3,IF(D{cur}="R2",$B$4,IF(D{cur}="R3",$B$5,"")))'
            ws.cell(cur, 6, assign)
            ws.cell(cur, 7, "未邀请")
            ws.cell(cur, 11, "Pending")
            cur += 1
    last = cur - 1
    if last >= hr + 1:
        style_data_rows(ws, hr + 1, last, len(headers))
        add_list_validation(ws, f"F{hr+1}:F{last}", YES_NO)
        add_list_validation(ws, f"G{hr+1}:G{last}", LOR_STATUS)
        add_table(ws, f"A{hr}:{get_column_letter(len(headers))}{last}", "tblLoR")
    ws.freeze_panes = f"A{hr+1}"


def build_essay_sheet(wb, projects, meta_map):
    ws = wb.create_sheet("04_Essay进度")
    headers = ["项目ID", "学校", "文书名称", "字数要求", "版本", "状态", "当前字数", "截止日期",
               "定稿日期", "本地文件路径", "协作链接(Google Doc)", "定制要点", "备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for p in projects:
        for essay_name, word_limit, local in meta_map[p["id"]].essays:
            ws.append([p["id"], p["school_zh"], essay_name, word_limit, "V0大纲", "未开始", "",
                       p["priority_ddl"], "", local, "", "", ""])
    last = ws.max_row
    style_data_rows(ws, 2, last, len(headers))
    add_list_validation(ws, f"E2:E{last}", ESSAY_VERSION)
    add_list_validation(ws, f"F2:F{last}", ESSAY_STATUS)
    ws.freeze_panes = "A2"
    add_table(ws, f"A1:{get_column_letter(len(headers))}{last}", "tblEssays")


def build_account_sheet(wb, projects, meta_map):
    ws = wb.create_sheet("05_网申账号")
    headers = ["项目ID", "学校", "项目名称", "网申系统", "Portal URL", "注册邮箱", "Application ID",
               "密码管理方式", "2FA方式", "注册状态", "账户备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for p in projects:
        m = meta_map[p["id"]]
        ws.append([p["id"], p["school_zh"], p["program"], m.portal_system, m.portal_url,
                   "", "", "密码管理器（不写明文）", "", "未注册", m.account_note])
    last = len(projects) + 1
    style_data_rows(ws, 2, last, len(headers))
    add_list_validation(ws, f"J2:J{last}", REG_STATUS)
    ws.freeze_panes = "A2"
    add_table(ws, f"A1:{get_column_letter(len(headers))}{last}", "tblAccounts")


def build_admission_sheet(wb, projects):
    ws = wb.create_sheet("06_录取进度")
    headers = ["项目ID", "学校", "项目名称", "提交日期", "材料齐全确认日", "审理状态", "结果通知日期",
               "决定类型", "回复截止日期", "是否接受", "押金截止", "备注"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for p in projects:
        ws.append([p["id"], p["school_zh"], p["program"], "", "", "已提交", "", "", "", "", "", ""])
    last = len(projects) + 1
    style_data_rows(ws, 2, last, len(headers))
    add_list_validation(ws, f"F2:F{last}", ADMISSION_STATUS)
    add_list_validation(ws, f"J2:J{last}", ["", "是", "否", "待定"])
    ws.freeze_panes = "A2"
    add_table(ws, f"A1:{get_column_letter(len(headers))}{last}", "tblAdmission")


def build_dashboard(wb, cfg, projects, meta_map):
    ws = wb.create_sheet("00_仪表盘", 0)
    ws["A1"] = f"{cfg.student_name} · 申请进度仪表盘"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:H1")
    ws["A3"], ws["B3"] = "学生姓名", cfg.student_name
    ws["C3"], ws["D3"] = "顾问", cfg.advisor_name
    ws["E3"], ws["F3"] = "生成日期", datetime.now().strftime("%Y-%m-%d")
    ws["G3"], ws["H3"] = "最近 Priority DDL", f"=MIN('01_项目总览'!G2:G{len(projects)+1})"
    ws["I3"], ws["J3"] = "距最近 DDL(天)", '=IF(H3="","",H3-TODAY())'
    for c in ["A3", "C3", "E3", "G3", "I3"]:
        ws[c].font = Font(bold=True)
    sr = 5
    ws.cell(sr, 1, "项目进度汇总").font = Font(bold=True, size=12)
    headers = ["项目ID", "学校", "Priority DDL", "距DDL(天)", "网申状态", "材料进度", "推荐信进度",
               "Essay进度", "整体就绪度(%)", "录取状态"]
    for c, h in enumerate(headers, 1):
        ws.cell(sr + 1, c, h)
    style_header_row(ws, sr + 1, len(headers))
    ps, adm = "'01_项目总览'", "'06_录取进度'"
    first = sr + 2
    for i in range(len(projects)):
        r, src = first + i, i + 2
        ws.cell(r, 1, f"={ps}!A{src}")
        ws.cell(r, 2, f"={ps}!B{src}")
        ws.cell(r, 3, f"={ps}!G{src}")
        ws.cell(r, 4, f'=IF(C{r}="","",C{r}-TODAY())')
        for j, col in enumerate([5, 6, 7, 8, 9, 10], 5):
            src_col = {5: "K", 6: "Q", 7: "R", 8: "S", 9: "U", 10: "F"}[j]
            sheet = ps if j < 10 else adm
            ws.cell(r, j, f"={sheet}!{src_col}{src}")
    last_s = first + len(projects) - 1
    style_data_rows(ws, first, last_s, len(headers))
    tr = last_s + 3
    ws.cell(tr, 1, "本周待办（距 Priority DDL ≤14 天且未 Complete）").font = Font(bold=True, size=12)
    th = ["项目ID", "学校", "Priority DDL", "距DDL(天)", "网申状态", "整体就绪度(%)", "当前卡点", "下一步动作", "负责人"]
    for c, h in enumerate(th, 1):
        ws.cell(tr + 1, c, h)
    style_header_row(ws, tr + 1, len(th))
    tf = tr + 2
    for i in range(len(projects)):
        r, src = tf + i, i + 2
        ws.cell(r, 1, f'=IF(AND({ps}!G{src}<>"",{ps}!G{src}-TODAY()<=14,{ps}!K{src}<>"Complete"),{ps}!A{src},"")')
        ws.cell(r, 2, f'=IF(A{r}="","",{ps}!B{src})')
        ws.cell(r, 3, f'=IF(A{r}="","",{ps}!G{src})')
        ws.cell(r, 4, f'=IF(A{r}="","",C{r}-TODAY())')
        ws.cell(r, 5, f'=IF(A{r}="","",{ps}!K{src})')
        ws.cell(r, 6, f'=IF(A{r}="","",{ps}!U{src})')
        ws.cell(r, 7, f'=IF(A{r}="","",{ps}!V{src})')
        ws.cell(r, 8, f'=IF(A{r}="","",{ps}!W{src})')
        ws.cell(r, 9, f'=IF(A{r}="","",{ps}!X{src})')
    style_data_rows(ws, tf, tf + len(projects) - 1, len(th))
    ws.freeze_panes = "A6"
    ws.conditional_formatting.add(f"D{first}:D{last_s}", CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=RED_FILL))
    ws.conditional_formatting.add(f"E{first}:E{last_s}", CellIsRule(operator="equal", formula=['"Complete"'], fill=GREEN_FILL))


def fix_overall_formula(wb, projects, meta_map, weights):
    ws = wb["01_项目总览"]
    w_def = weights.get("default", {"material": 0.4, "lor": 0.3, "essay": 0.2, "app": 0.1})
    w_nl = weights.get("no_lor", {"material": 0.4, "lor": 0.0, "essay": 0.2, "app": 0.1})
    for r in range(2, len(projects) + 2):
        pid = ws.cell(r, 1).value
        w = w_nl if meta_map[pid].lor_count == 0 else w_def
        parts = [
            f"IFERROR(VALUE(LEFT(Q{r},FIND(\"/\",Q{r}&\"/\")-1))/IFERROR(VALUE(MID(Q{r},FIND(\"/\",Q{r}&\"/\")+1,99)),1),0)*{w['material']}",
        ]
        if w["lor"] > 0:
            parts.append(
                f"IFERROR(VALUE(LEFT(R{r},FIND(\"/\",R{r}&\"/\")-1))/IFERROR(VALUE(MID(R{r},FIND(\"/\",R{r}&\"/\")+1,99)),1),0)*{w['lor']}"
            )
        parts.append(
            f"IFERROR(VALUE(LEFT(S{r},FIND(\"/\",S{r}&\"/\")-1))/IFERROR(VALUE(MID(S{r},FIND(\"/\",S{r}&\"/\")+1,99)),1),0)*{w['essay']}"
        )
        parts.append(f"T{r}*{w['app']}")
        ws.cell(r, 21).value = f"=ROUND({'+'.join(parts)},0)"


def apply_conditional_formatting(wb, n):
    ws = wb["01_项目总览"]
    last = n + 1
    ws.conditional_formatting.add(f"G2:G{last}", FormulaRule(formula=["G2-TODAY()<=7"], fill=YELLOW_FILL))
    ws.conditional_formatting.add(f"G2:G{last}", FormulaRule(formula=["G2-TODAY()<0"], fill=RED_FILL))
    ws.conditional_formatting.add(f"K2:K{last}", CellIsRule(operator="equal", formula=['"Complete"'], fill=GREEN_FILL))
    mat = wb["02_材料进度"]
    ml = mat.max_row
    mat.conditional_formatting.add(f"E3:E{ml}", CellIsRule(operator="equal", formula=['"已上传"'], fill=GREEN_FILL))
    lor = wb["03_推荐信进度"]
    if lor.max_row >= 9:
        lor.conditional_formatting.add(f"G9:G{lor.max_row}", CellIsRule(operator="equal", formula=['"已提交"'], fill=GREEN_FILL))
    essay = wb["04_Essay进度"]
    el = essay.max_row
    essay.conditional_formatting.add(f"F2:F{el}", CellIsRule(operator="equal", formula=['"已定稿"'], fill=GREEN_FILL))


def write_readme(cfg, projects, output_dir, skill_dir):
    lines = [
        f"# {cfg.student_name} · 申请进度 Checklist 使用说明",
        "",
        f"文件：[{cfg.output_filename}](./{cfg.output_filename})",
        "",
        f"共 {len(projects)} 个项目。通用材料（项目ID=`通用`）只需维护一次。",
        "",
        "Google Sheets 协作请参阅 skill 文档：`sheets-guide.md`",
        "",
        f"重新生成：`python3 {skill_dir}/scripts/build_checklist.py --config checklist.config.yaml [--merge 现有.xlsx]`",
    ]
    (output_dir / f"{cfg.student_name}_申请进度Checklist_使用说明.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Build application progress checklist")
    parser.add_argument("--config", required=True, help="Path to checklist.config.yaml")
    parser.add_argument("--merge", help="Existing xlsx to merge user progress from")
    args = parser.parse_args()

    cfg = load_config(Path(args.config).resolve())
    projects, meta_map = parse_projects(cfg)
    if not projects:
        raise SystemExit("No projects found in school list CSV")

    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    out_path = cfg.output_dir / cfg.output_filename
    wb = build_workbook(cfg, projects, meta_map)
    wb.save(out_path)

    if args.merge:
        stats = merge_workbook(out_path, Path(args.merge).resolve())
        print(f"Merged progress from {args.merge}: {stats}")

    write_readme(cfg, projects, cfg.output_dir, SKILL_DIR)
    print(f"Created: {out_path}")
    print(f"Projects ({len(projects)}): {[p['id'] for p in projects]}")


if __name__ == "__main__":
    main()
